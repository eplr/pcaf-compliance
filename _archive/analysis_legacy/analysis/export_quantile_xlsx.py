#!/usr/bin/env python3
"""
Export PCAF Quantile-Based Thresholds to Excel
Format identique à partial_coverage_analysis_EN.xlsx
"""

import csv
from pathlib import Path
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = Path("/Users/fidestra/Library/CloudStorage/OneDrive-fidestra/03_ACCOUNTS/Axylia/pcaf-compliance/analysis/results")
OUTPUT_PATH = RESULTS_DIR / "pcaf_quantile_analysis_EN.xlsx"

# ── Couleurs (identiques au fichier de référence) ──
FILL_REPORTED = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # vert
FILL_PARTIAL  = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # jaune
FILL_MISSING  = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # rouge
FILL_NA       = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # gris
FILL_HEADER   = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")  # gris foncé
FILL_NONE     = None

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_TITLE  = Font(bold=True, size=12)
FONT_SUBTITLE = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_LEGEND = Font(bold=True, size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

STATUS_SYMBOL = {"REPORTED": "✓", "PARTIAL": "◐", "MISSING": "✗", "N/A": "—"}
STATUS_FILL   = {"REPORTED": FILL_REPORTED, "PARTIAL": FILL_PARTIAL, "MISSING": FILL_MISSING, "N/A": FILL_NA}

# Colonnes Part A (10 asset classes)
PART_A_COLS = [
    "Listed equity and corporate bonds",
    "Business loans and unlisted equity",
    "Project finance",
    "Commercial real estate",
    "Mortgages",
    "Motor vehicle loans",
    "Use of proceeds",
    "Securitized and structured products",
    "Sovereign debt",
    "Sub-sovereign debt",
]

# Colonnes Part B
PART_B_COLS = ["Facilitated emissions"]

# Colonnes Part C
PART_C_COLS = ["Commercial lines", "Personal motor", "Project insurance", "Treaty reinsurance"]

# En-têtes abrégés pour la matrice
SHORT_HEADERS = {
    "Listed equity and corporate bonds": "Listed Eq. &\nCorp. Bonds",
    "Business loans and unlisted equity": "Business Loans /\nUnlisted Eq.",
    "Project finance": "Project\nFinance",
    "Commercial real estate": "Commercial\nReal Estate",
    "Mortgages": "Residential\nMortgages",
    "Motor vehicle loans": "Motor Vehicle\nLoans",
    "Use of proceeds": "Use of\nProceeds",
    "Securitized and structured products": "Securitized /\nStructured",
    "Sovereign debt": "Sovereign\nDebt",
    "Sub-sovereign debt": "Sub-Sovereign\nDebt",
    "Facilitated emissions": "Facilitated\nEmissions",
    "Commercial lines": "Commercial\nLines",
    "Personal motor": "Personal\nMotor",
    "Project insurance": "Project\nInsurance",
    "Treaty reinsurance": "Treaty\nReinsurance",
}

PCAF_REQUIREMENTS = {
    "insurer":       {"part_a": True, "part_b": False, "part_c": True},
    "reinsurer":     {"part_a": True, "part_b": False, "part_c": True},
    "bank":          {"part_a": True, "part_b": True,  "part_c": False},
    "asset_manager": {"part_a": True, "part_b": False, "part_c": False},
    "bancassurance": {"part_a": True, "part_b": True,  "part_c": True},
    "other":         {"part_a": True, "part_b": False, "part_c": False},
}


def load_csv(filename):
    path = RESULTS_DIR / filename
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def set_cell(ws, row, col, value, font=FONT_NORMAL, fill=FILL_NONE, alignment=ALIGN_CENTER, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def build_classification_sheet(wb, classification, thresholds):
    """Sheet 1: Classification Matrix (style identique à Asset Class Matrix)"""
    ws = wb.active
    ws.title = "Classification Matrix"

    ALL_CATS = PART_A_COLS + PART_B_COLS + PART_C_COLS
    n_cats = len(ALL_CATS)

    # ── Row 1: Titre ──
    set_cell(ws, 1, 1, "PCAF Quantile-Based Classification Matrix", font=FONT_TITLE, alignment=ALIGN_LEFT)

    # ── Row 3: Légende ──
    set_cell(ws, 3, 1, "Legend:", font=FONT_LEGEND, alignment=ALIGN_LEFT)
    set_cell(ws, 3, 2, "Reported", font=FONT_LEGEND, fill=FILL_REPORTED)
    set_cell(ws, 3, 3, "Partial", font=FONT_LEGEND, fill=FILL_PARTIAL)
    set_cell(ws, 3, 4, "Missing", font=FONT_LEGEND, fill=FILL_MISSING)
    set_cell(ws, 3, 5, "N/A", font=FONT_LEGEND, fill=FILL_NA)

    # ── Row 5: En-têtes ──
    header_row = 5
    headers = ["Institution", "Type"] + [SHORT_HEADERS.get(c, c) for c in ALL_CATS] + ["Coverage"]
    for col_idx, h in enumerate(headers, 1):
        set_cell(ws, header_row, col_idx, h, font=FONT_HEADER, fill=FILL_HEADER, border=THIN_BORDER)

    # ── Construire les données par institution ──
    institutions = {}
    for row in classification:
        inst = row["institution"]
        if inst not in institutions:
            institutions[inst] = {"type": row["institution_type"], "cats": {}}
        institutions[inst]["cats"][(row["part"], row["category"])] = row["status"]

    # Calculer la couverture (% non-MISSING parmi les catégories applicables)
    inst_coverage = {}
    for inst, data in institutions.items():
        reqs = PCAF_REQUIREMENTS.get(data["type"], {"part_a": True, "part_b": False, "part_c": False})
        applicable = 0
        covered = 0
        for cat in ALL_CATS:
            part = "Part A" if cat in PART_A_COLS else ("Part B" if cat in PART_B_COLS else "Part C")
            part_key = part.split(" ")[1].lower()
            req_key = f"part_{part_key}"
            if not reqs.get(req_key, False):
                continue
            applicable += 1
            status = data["cats"].get((part, cat), "MISSING")
            if status != "MISSING":
                covered += 1
        inst_coverage[inst] = covered / applicable if applicable > 0 else 0

    # Trier par couverture décroissante
    sorted_insts = sorted(institutions.keys(), key=lambda x: -inst_coverage[x])

    # Groupes
    groups = [
        ("High Coverage (≥ 80%)", lambda c: c >= 0.80),
        ("Medium Coverage (60–80%)", lambda c: 0.60 <= c < 0.80),
        ("Low Coverage (< 60%)", lambda c: c < 0.60),
    ]

    current_row = header_row + 1
    for group_label, group_filter in groups:
        group_insts = [i for i in sorted_insts if group_filter(inst_coverage[i])]
        if not group_insts:
            continue

        # Sous-titre du groupe
        set_cell(ws, current_row, 1, group_label, font=FONT_SUBTITLE, alignment=ALIGN_LEFT)
        current_row += 1

        for inst in group_insts:
            data = institutions[inst]
            reqs = PCAF_REQUIREMENTS.get(data["type"], {"part_a": True, "part_b": False, "part_c": False})

            set_cell(ws, current_row, 1, inst, font=FONT_NORMAL, alignment=ALIGN_LEFT, border=THIN_BORDER)
            set_cell(ws, current_row, 2, data["type"].replace("_", " ").title(), font=FONT_NORMAL, alignment=ALIGN_CENTER, border=THIN_BORDER)

            for cat_idx, cat in enumerate(ALL_CATS):
                col = cat_idx + 3
                part = "Part A" if cat in PART_A_COLS else ("Part B" if cat in PART_B_COLS else "Part C")
                part_key = part.split(" ")[1].lower()
                req_key = f"part_{part_key}"

                if not reqs.get(req_key, False):
                    status = "N/A"
                else:
                    status = data["cats"].get((part, cat), "MISSING")

                symbol = STATUS_SYMBOL.get(status, "?")
                fill = STATUS_FILL.get(status, FILL_NONE)
                set_cell(ws, current_row, col, symbol, fill=fill, border=THIN_BORDER)

            # Coverage
            cov = inst_coverage[inst]
            set_cell(ws, current_row, len(ALL_CATS) + 3, round(cov, 4), font=FONT_NORMAL, border=THIN_BORDER)

            current_row += 1

        # Ligne vide entre groupes
        current_row += 1

    # ── Largeurs de colonnes ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for i in range(3, len(ALL_CATS) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[header_row].height = 35


def build_composite_scores_sheet(wb, scores):
    """Sheet 2: Scores composites détaillés"""
    ws = wb.create_sheet("Composite Scores")

    set_cell(ws, 1, 1, "Composite Scores (0–12) per Institution × Category", font=FONT_TITLE, alignment=ALIGN_LEFT)
    set_cell(ws, 2, 1, "Dimensions: Mention (0-3) + Quantitative (0-3) + Coverage (0-3) + Data Quality (0-3)", font=FONT_NORMAL, alignment=ALIGN_LEFT)

    ALL_CATS = PART_A_COLS + PART_B_COLS + PART_C_COLS

    # En-têtes
    header_row = 4
    headers = ["Institution", "Type"] + [SHORT_HEADERS.get(c, c) for c in ALL_CATS]
    for col_idx, h in enumerate(headers, 1):
        set_cell(ws, header_row, col_idx, h, font=FONT_HEADER, fill=FILL_HEADER, border=THIN_BORDER)

    # Données par institution
    institutions = {}
    for row in scores:
        inst = row["institution"]
        if inst not in institutions:
            institutions[inst] = {"type": row["institution_type"], "cats": {}}
        institutions[inst]["cats"][(row["part"], row["category"])] = int(row["composite"])

    current_row = header_row + 1
    for inst in sorted(institutions.keys()):
        data = institutions[inst]
        set_cell(ws, current_row, 1, inst, font=FONT_NORMAL, alignment=ALIGN_LEFT, border=THIN_BORDER)
        set_cell(ws, current_row, 2, data["type"].replace("_", " ").title(), font=FONT_NORMAL, border=THIN_BORDER)

        for cat_idx, cat in enumerate(ALL_CATS):
            col = cat_idx + 3
            part = "Part A" if cat in PART_A_COLS else ("Part B" if cat in PART_B_COLS else "Part C")
            composite = data["cats"].get((part, cat), 0)

            # Couleur dégradée selon le score
            if composite == 0:
                fill = FILL_MISSING
            elif composite <= 2:
                fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif composite <= 4:
                fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # orange clair
            elif composite <= 6:
                fill = FILL_PARTIAL
            elif composite <= 8:
                fill = PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")  # vert clair
            else:
                fill = FILL_REPORTED

            set_cell(ws, current_row, col, composite, fill=fill, border=THIN_BORDER)

        current_row += 1

    # Légende
    current_row += 2
    set_cell(ws, current_row, 1, "Score Legend:", font=FONT_LEGEND, alignment=ALIGN_LEFT)
    legend_items = [
        (0, "0 (Not found)", FILL_MISSING),
        (2, "1–2 (Minimal)", PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")),
        (4, "3–4 (Basic)", PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")),
        (6, "5–6 (Moderate)", FILL_PARTIAL),
        (8, "7–8 (Good)", PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")),
        (10, "9–12 (Comprehensive)", FILL_REPORTED),
    ]
    for i, (_, label, fill) in enumerate(legend_items):
        set_cell(ws, current_row, 2 + i, label, font=FONT_LEGEND, fill=fill)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for i in range(3, len(ALL_CATS) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[header_row].height = 35


def build_thresholds_sheet(wb, thresholds):
    """Sheet 3: Seuils par quantiles"""
    ws = wb.create_sheet("Quantile Thresholds")

    set_cell(ws, 1, 1, "Empirical Thresholds by Category (based on panel quantiles)", font=FONT_TITLE, alignment=ALIGN_LEFT)

    header_row = 3
    headers = ["Part", "Category", "N", "Min", "Q1", "Median", "Q3", "Max", "Mean",
               "Threshold PARTIAL", "Threshold REPORTED"]
    for col_idx, h in enumerate(headers, 1):
        set_cell(ws, header_row, col_idx, h, font=FONT_HEADER, fill=FILL_HEADER, border=THIN_BORDER)

    for i, row in enumerate(thresholds, header_row + 1):
        vals = [row["part"], row["category"], row["n"], row["min"], row["q1"],
                row["median"], row["q3"], row["max"], row["mean"],
                row["seuil_partial"], row["seuil_reported"]]
        for col_idx, v in enumerate(vals, 1):
            try:
                v = float(v)
            except (ValueError, TypeError):
                pass
            set_cell(ws, i, col_idx, v, border=THIN_BORDER,
                     alignment=ALIGN_CENTER if col_idx >= 3 else ALIGN_LEFT)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    for i in range(3, 12):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[header_row].height = 30


def build_summary_sheet(wb, summary):
    """Sheet 4: Résumé par institution × Part"""
    ws = wb.create_sheet("Institution Summary")

    set_cell(ws, 1, 1, "PCAF Compliance Summary by Institution (Quantile-Based)", font=FONT_TITLE, alignment=ALIGN_LEFT)

    header_row = 3
    headers = ["Institution", "Type", "PCAF Part", "Status", "Avg Score", "Max Score",
               "Reported", "Partial", "Missing", "Total", "Completeness %", "Missing Elements"]
    for col_idx, h in enumerate(headers, 1):
        set_cell(ws, header_row, col_idx, h, font=FONT_HEADER, fill=FILL_HEADER, border=THIN_BORDER)

    for i, row in enumerate(summary, header_row + 1):
        vals = [
            row["institution"], row["institution_type"], row["part"], row["status"],
            float(row["avg_composite_score"]), int(row["max_composite_score"]),
            int(row["reported_count"]), int(row["partial_count"]),
            int(row["missing_count"]), int(row["total_categories"]),
            float(row["completeness_pct"]), row["missing_elements"]
        ]
        for col_idx, v in enumerate(vals, 1):
            fill = FILL_NONE
            if col_idx == 4:  # Status column
                fill = STATUS_FILL.get(v, FILL_NONE)
            align = ALIGN_LEFT if col_idx in (1, 2, 3, 12) else ALIGN_CENTER
            set_cell(ws, i, col_idx, v, fill=fill, border=THIN_BORDER, alignment=align)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    for i in range(5, 12):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["L"].width = 50
    ws.row_dimensions[header_row].height = 30


def main():
    print("Loading CSV results...")
    classification = load_csv("pcaf_quantile_classification.csv")
    scores = load_csv("pcaf_composite_scores.csv")
    thresholds = load_csv("pcaf_quantile_thresholds.csv")
    summary = load_csv("pcaf_quantile_summary.csv")

    wb = Workbook()

    print("Building Classification Matrix...")
    build_classification_sheet(wb, classification, thresholds)

    print("Building Composite Scores...")
    build_composite_scores_sheet(wb, scores)

    print("Building Quantile Thresholds...")
    build_thresholds_sheet(wb, thresholds)

    print("Building Institution Summary...")
    build_summary_sheet(wb, summary)

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Excel exporté : {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
