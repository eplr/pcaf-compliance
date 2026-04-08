#!/usr/bin/env python3
"""Translate partial_coverage_analysis_EN.xlsx to French using TMX translation memory."""

import re
import copy
from openpyxl import load_workbook

SRC = "/Users/fidestra/Library/CloudStorage/OneDrive-fidestra/03_ACCOUNTS/Axylia/pcaf-compliance/translation/partial_coverage_analysis_EN.xlsx"
DST = "/Users/fidestra/Library/CloudStorage/OneDrive-fidestra/03_ACCOUNTS/Axylia/pcaf-compliance/translation/partial_coverage_analysis_FR.xlsx"

# ── TMX-based translation dictionary ──────────────────────────────────────────

# Institution types (capitalized for display)
TYPE_MAP = {
    "Insurer": "Assurance",
    "Bank": "Banque",
    "Bancassurance": "Bancassurance",
    "Reinsurer": "Réassurance",
    "Asset Manager": "Gestionnaire d'actifs",
    "Exchange": "Opérateur de bourse",
    "Inv. Holding": "Holding d'invest.",
}

# Asset class names (short forms used in columns I/J/K of Summary)
# NB: Sub-Sovereign Debt MUST come before Sovereign Debt to avoid partial match
ASSET_CLASS_MAP = {
    "Listed Eq. & Corp. Bonds": "Actions et obligations cotées",
    "Business Loans / Unlisted Eq.": "Prêts aux entreprises / capitaux propres non cotés",
    "Project Finance": "Financement de projets",
    "Commercial Real Estate": "Immobilier commercial",
    "Residential Mortgages": "Hypothèques",
    "Motor Vehicle Loans": "Véhicules à moteur",
    "Use of Proceeds": "Utilisation des produits",
    "Securitized / Structured": "Produits titrisés / structurés",
    "Sub-Sovereign Debt": "Dette sub-souveraine",
    "Sovereign Debt": "Dette souveraine",
    "Capital Markets": "Marchés des capitaux",
    "Advisory Services": "Services de conseil",
    "Insurance Underwriting": "Souscription d'assurance",
}

# Matrix column headers (with newlines)
MATRIX_HEADER_MAP = {
    "Listed Eq. &\nCorp. Bonds": "Actions et\nobl. cotées",
    "Business Loans /\nUnlisted Eq.": "Prêts aux entr. /\ncap. pr. non cotés",
    "Project\nFinance": "Financement\nde projets",
    "Commercial\nReal Estate": "Immobilier\ncommercial",
    "Residential\nMortgages": "Hypothèques",
    "Motor Vehicle\nLoans": "Véhicules\nà moteur",
    "Use of\nProceeds": "Utilisation\ndes produits",
    "Securitized /\nStructured": "Produits titrisés\n/ structurés",
    "Sovereign\nDebt": "Dette\nsouveraine",
    "Sub-Sovereign\nDebt": "Dette\nsub-souveraine",
    "Capital\nMarkets": "Marchés des\ncapitaux",
    "Advisory\nServices": "Services\nde conseil",
    "Insurance\nUnderwriting": "Souscription\nd'assurance",
}

# Part status terms
STATUS_MAP = {
    "reported": "reporting réalisé",
    "partial": "partiel",
    "missing": "manquant",
}

# Generic term translations
TERM_MAP = {
    "Institution": "Institution",
    "Type": "Type",
    "Part Statuses": "Statuts des parties",
    "Partial\nParts": "Parties\npartielles",
    "Reported": "Reporting réalisé",
    "Partial": "Partiel",
    "Missing": "Manquant",
    "Coverage": "Couverture",
    "Reported Asset Classes": "Classes d'actifs –\nreporting réalisé",
    "Partial Asset Classes": "Classes d'actifs –\nreporting partiel",
    "Missing Asset Classes": "Classes d'actifs –\nmanquantes",
    "Legend:": "Légende :",
    "N/A": "N/A",
}


def translate_part_statuses(text):
    """Translate 'A: reported | B: partial' style strings."""
    if not isinstance(text, str):
        return text
    parts = text.split(" | ")
    translated = []
    for p in parts:
        m = re.match(r"^([ABC]):\s*(\w+)$", p.strip())
        if m:
            letter, status = m.group(1), m.group(2)
            fr_status = STATUS_MAP.get(status, status)
            translated.append(f"{letter} : {fr_status}")
        else:
            translated.append(p)
    return " | ".join(translated)


def translate_asset_class_list(text):
    """Translate '[A] Listed Eq. & Corp. Bonds, [B] Capital Markets' style strings."""
    if not isinstance(text, str) or text == "—":
        return text
    for en, fr in ASSET_CLASS_MAP.items():
        text = text.replace(en, fr)
    return text


def translate_summary_title(text):
    """Translate the main title of the Summary sheet."""
    if text == "PCAF Asset Class Coverage Analysis — All Institutions":
        return "Analyse de couverture des classes d'actifs PCAF® – Tous les établissements"
    return text


def translate_coverage_tiers(text):
    """Translate coverage tier headers."""
    tier_map = {
        "High Coverage (≥ 80%)": "Couverture élevée (≥ 80 %)",
        "Medium Coverage (60–80%)": "Couverture moyenne (60–80 %)",
        "Low Coverage (< 60%)": "Couverture faible (< 60 %)",
    }
    return tier_map.get(text, text)


def translate_description_rows(text):
    """Translate the description rows (rows 3–9) of the Summary sheet."""
    desc_map = {
        "Applicable asset classes by institution type:":
            "Classes d'actifs applicables par type d'institution :",
        "  • Bancassurance: 13/13 (all classes applicable)":
            "  • Bancassurance : 13/13 (toutes les classes applicables)",
        "  • Banks: 12/13 (excl. Insurance Underwriting)":
            "  • Banques : 12/13 (hors Souscription d'assurance)",
        "  • Insurers / Reinsurers: 11/13 (excl. Capital Markets, Advisory Services)":
            "  • Assurance / Réassurance : 11/13 (hors Marchés des capitaux, Services de conseil)",
        "  • Asset Managers: 10/13 (excl. Capital Markets, Advisory Services, Insurance Underwriting)":
            "  • Gestionnaires d'actifs : 10/13 (hors Marchés des capitaux, Services de conseil, Souscription d'assurance)",
        "  • Exchange / Inv. Holding: 10/13 (excl. Capital Markets, Advisory Services, Insurance Underwriting)":
            "  • Opérateurs de bourse / Holdings d'invest. : 10/13 (hors Marchés des capitaux, Services de conseil, Souscription d'assurance)",
        "[A] = Part A (Financed Emissions)  |  [B] = Part B (Facilitated Emissions)  |  [C] = Part C (Insurance-Associated Emissions)":
            "[A] = Partie A (émissions financées)  |  [B] = Partie B (émissions facilitées)  |  [C] = Partie C (émissions liées à l'assurance)",
    }
    return desc_map.get(text, text)


def translate_matrix_title(text):
    """Translate the main title of the Asset Class Matrix sheet."""
    if text == "Detailed Asset Class Coverage Matrix — All Institutions":
        return "Matrice détaillée de couverture des classes d'actifs – Tous les établissements"
    return text


def translate_matrix_part_headers(text):
    """Translate Part headers in the matrix."""
    part_map = {
        "Part A — Financed Emissions": "Partie A – émissions financées",
        "Part B": "Partie B",
        "Part C": "Partie C",
    }
    return part_map.get(text, text)


def translate_footer(text):
    """Translate the footer formula explanation."""
    if text == "Coverage = (Reported + Partial) / Applicable asset classes":
        return "Couverture = (reporting réalisé + partiel) / classes d'actifs applicables"
    return text


# ── Main processing ───────────────────────────────────────────────────────────

wb = load_workbook(SRC)

# Process Summary sheet
ws = wb["Summary"]
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if v is None or not isinstance(v, str):
            continue
        col = cell.column  # 1-based
        row_num = cell.row

        # Row 1: main title
        if row_num == 1:
            cell.value = translate_summary_title(v)
            continue

        # Rows 3–9: description rows
        if 3 <= row_num <= 9:
            cell.value = translate_description_rows(v)
            continue

        # Coverage tier headers (rows 11, 23, 34)
        tier = translate_coverage_tiers(v)
        if tier != v:
            cell.value = tier
            continue

        # Column header rows (rows 12, 24, 35)
        if row_num in (12, 24, 35):
            if v in TERM_MAP:
                cell.value = TERM_MAP[v]
            continue

        # Footer row
        if row_num == 43:
            cell.value = translate_footer(v)
            continue

        # Data rows: translate by column
        # Col B: institution type
        if col == 2 and v in TYPE_MAP:
            cell.value = TYPE_MAP[v]
            continue

        # Col C: Part statuses
        if col == 3:
            cell.value = translate_part_statuses(v)
            continue

        # Cols I, J, K (9, 10, 11): asset class lists
        if col in (9, 10, 11):
            cell.value = translate_asset_class_list(v)
            continue

# Process Asset Class Matrix sheet
ws2 = wb["Asset Class Matrix"]
for row in ws2.iter_rows():
    for cell in row:
        v = cell.value
        if v is None or not isinstance(v, str):
            continue
        col = cell.column
        row_num = cell.row

        # Row 1: title
        if row_num == 1:
            cell.value = translate_matrix_title(v)
            continue

        # Row 3: legend + status labels
        if row_num == 3:
            legend_map = {
                "Legend:": "Légende :",
                "Reported": "Reporting réalisé",
                "Partial": "Partiel",
                "Missing": "Manquant",
                "N/A": "N/A",
            }
            if v in legend_map:
                cell.value = legend_map[v]
            continue

        # Row 4: Part headers
        if row_num == 4:
            cell.value = translate_matrix_part_headers(v)
            continue

        # Row 5: column headers
        if row_num == 5:
            if v in MATRIX_HEADER_MAP:
                cell.value = MATRIX_HEADER_MAP[v]
            elif v in TERM_MAP:
                cell.value = TERM_MAP[v]
            continue

        # Coverage tier separator rows
        tier = translate_coverage_tiers(v)
        if tier != v:
            cell.value = tier
            continue

        # Col B: institution type
        if col == 2 and v in TYPE_MAP:
            cell.value = TYPE_MAP[v]
            continue

        # Col C: Part statuses
        if col == 3:
            cell.value = translate_part_statuses(v)
            continue

# Rename sheets
ws.title = "Synthèse"
ws2.title = "Matrice de couverture"

wb.save(DST)
print(f"French version saved to:\n{DST}")
